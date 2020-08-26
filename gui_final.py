# -*- coding: utf-8 -*-
"""
Created on Tue Mar 26 19:11:25 2019

@author: Bijal Patel
"""

# import openpyxl and tkinter modules 
from openpyxl import *
import tkinter as tk
import pandas as pd
import numpy as np

# globally declare wb and sheet variable 

# opening the existing excel file 
wb = load_workbook('D:\Studies\Engineering\Major Project\Thyroid\SEM 8\Python - Sypder\excel.xlsx') 

# create the sheet object 
sheet = wb.active 

def lda():
    # Importing the dataset
    dataset = pd.read_csv('D:/Major Project/Thyroid/SEM 8/Python - Sypder/hyper_dataset1.csv')
    X = dataset.iloc[:, 0:20].values
    y = dataset.iloc[:, 20].values
    dataset = pd.read_csv('D:/Major Project/Thyroid/SEM 8/Python - Sypder/excel.csv')
    X_test = dataset.iloc[:, 0:20].values
    
    # Taking care of missing data
    from sklearn.impute import SimpleImputer
    imputer = SimpleImputer(missing_values = np.nan, strategy = 'median')
    imputer = imputer.fit(X[:, 15:20])
    X[:, 15:20] = imputer.transform(X[:, 15:20])
    imputer = imputer.fit(X[:,1:2])
    X[:,1:2] = imputer.transform(X[:,1:2])

    # Splitting the dataset into the Training set and Test set
    #from sklearn.model_selection import train_test_split
    #X_train, X_test, y_train, y_test = train_test_split(X, y, test_size = 0.2, random_state = 0)

    # Feature Scaling
    from sklearn.preprocessing import StandardScaler
    sc = StandardScaler()
    X = sc.fit_transform(X)
    X_test = sc.transform(X_test)

    #Converting infinte to finite
    X = np.nan_to_num(X)

    # Applying LDA
    from sklearn.discriminant_analysis import LinearDiscriminantAnalysis as LDA
    lda = LDA()
    X = lda.fit_transform(X, y)
    X_test = lda.transform(X_test)

    # Fitting Logistic Regression to the Training set
    from sklearn.linear_model import LogisticRegression
    classifier = LogisticRegression(random_state = 0)
    classifier.fit(X, y)
    # Predicting the Test set results
    y_pred = classifier.predict(X_test)
    return y_pred[-1]

def popup(title, msg):
    '''Open popup window with title and msg'''
    w = tk.Toplevel(root)
    w.title(title)
    w.minsize(200, 200)
    tk.Label(w, text=msg).pack()
    tk.Button(w, text="Close", command=w.destroy).pack(pady=10)
    w.bind("<Return>", lambda f: w.destroy())


def excel(): 
	
	# resize the width of columns in 
   # excel spreadsheet 
   sheet.column_dimensions['A'].width = 10
   sheet.column_dimensions['B'].width = 10
   sheet.column_dimensions['C'].width = 10
   sheet.column_dimensions['D'].width = 10
   sheet.column_dimensions['E'].width = 10
   sheet.column_dimensions['F'].width = 10
   sheet.column_dimensions['G'].width = 10
   sheet.column_dimensions['H'].width = 10
   sheet.column_dimensions['I'].width = 10
   sheet.column_dimensions['J'].width = 10
   sheet.column_dimensions['K'].width = 10
   sheet.column_dimensions['L'].width = 10
   sheet.column_dimensions['M'].width = 10
   sheet.column_dimensions['N'].width = 10
   sheet.column_dimensions['O'].width = 10
   sheet.column_dimensions['P'].width = 10
   sheet.column_dimensions['Q'].width = 10
   sheet.column_dimensions['R'].width = 10
   sheet.column_dimensions['S'].width = 10
   sheet.column_dimensions['T'].width = 10
   

	# write given data to an excel spreadsheet 
	# at particular location 
   sheet.cell(row=1, column=1).value = "Age"
   sheet.cell(row=1, column=2).value = "Gender"
   sheet.cell(row=1, column=3).value = "On_Thyroxine"
   sheet.cell(row=1, column=4).value = "Query_on_Thyroxine"
   sheet.cell(row=1, column=5).value = "On_antithyroid_medication"
   sheet.cell(row=1, column=6).value = "Sick"
   sheet.cell(row=1, column=7).value = "Pregnant"
   sheet.cell(row=1, column=8).value = "Thyroid_Surgery"
   sheet.cell(row=1, column=9).value = "T131_treatment"
   sheet.cell(row=1, column=10).value = "Query_hypothyroid"
   sheet.cell(row=1, column=11).value = "Query_hyperthyroid"
   sheet.cell(row=1, column=12).value = "Lithium"
   sheet.cell(row=1, column=13).value = "Goitre"
   sheet.cell(row=1, column=14).value = "Tumor"
   sheet.cell(row=1, column=15).value = "Psych"
   sheet.cell(row=1, column=16).value = "TSH"
   sheet.cell(row=1, column=17).value = "T3"
   sheet.cell(row=1, column=18).value = "TT4"
   sheet.cell(row=1, column=19).value = "T4U"
   sheet.cell(row=1, column=20).value = "FTI"
    
# Function to set focus (cursor) 
def focus1(event): 
	# set focus on the course_field box 
	Gender_field.focus_set() 

# Function to set focus 
def focus2(event): 
	# set focus on the sem_field box 
	On_Thyroxine_field.focus_set() 

# Function to set focus 
def focus3(event): 
	# set focus on the form_no_field box 
	Query_on_Thyroxine_field.focus_set() 

# Function to set focus 
def focus4(event): 
	# set focus on the contact_no_field box 
	On_antithyroid_medication_field.focus_set() 

# Function to set focus 
def focus5(event): 
	# set focus on the email_id_field box 
	Sick_field.focus_set() 

# Function to set focus 
def focus6(event): 
	# set focus on the address_field box 
	Pregnant_field.focus_set() 

# Function to set focus 
def focus7(event): 
	# set focus on the address_field box 
	Thyroid_Surgery_field.focus_set() 

# Function to set focus 
def focus8(event): 
	# set focus on the address_field box 
	T131_treatment_field.focus_set() 

# Function to set focus 
def focus9(event): 
	# set focus on the address_field box 
	Query_hypothyroid_field.focus_set() 

# Function to set focus 
def focus10(event): 
	# set focus on the address_field box 
	Query_hyperthyroid_field.focus_set() 

# Function to set focus 
def focus11(event): 
	# set focus on the address_field box 
	Lithium_field.focus_set() 

# Function to set focus 
def focus12(event): 
	# set focus on the address_field box 
	Goitre_field.focus_set() 

# Function to set focus 
def focus13(event): 
	# set focus on the address_field box 
	Tumor_field.focus_set() 

# Function to set focus 
def focus14(event): 
	# set focus on the address_field box 
	Psych_field.focus_set() 

# Function to set focus 
def focus15(event): 
	# set focus on the address_field box 
	TSH_field.focus_set() 

# Function to set focus 
def focus16(event): 
	# set focus on the address_field box 
	T3_field.focus_set() 

# Function to set focus 
def focus17(event): 
	# set focus on the address_field box 
	TT4_field.focus_set() 

# Function to set focus 
def focus18(event): 
	# set focus on the address_field box 
	T4U_field.focus_set() 

# Function to set focus 
def focus19(event): 
	# set focus on the address_field box 
	FTI_field.focus_set() 

# Function for clearing the 
# contents of text entry boxes 
def clear(): 
	
	# clear the content of text entry box 
   Age_field.delete(0, tk.END) 
   Gender_field.delete(0, tk.END) 
   On_Thyroxine_field.delete(0, tk.END) 
   Query_on_Thyroxine_field.delete(0, tk.END) 
   On_antithyroid_medication_field.delete(0, tk.END) 
   Sick_field.delete(0, tk.END) 
   Pregnant_field.delete(0, tk.END)
   Thyroid_Surgery_field.delete(0, tk.END)
   T131_treatment_field.delete(0, tk.END)
   Query_hypothyroid_field.delete(0, tk.END)
   Query_hyperthyroid_field.delete(0, tk.END)
   Lithium_field.delete(0, tk.END)
   Goitre_field.delete(0, tk.END)
   Tumor_field.delete(0, tk.END)
   Psych_field.delete(0, tk.END)
   TSH_field.delete(0, tk.END)
   T3_field.delete(0, tk.END)
   TT4_field.delete(0, tk.END)
   T4U_field.delete(0, tk.END)
   FTI_field.delete(0, tk.END)


# Function to take data from GUI 
# window and write to an excel file 
def insert(): 
	
	# if user not fill any entry 
	# then print "empty input" 
   if (Age_field.get() == "" and
     Gender_field.get() == "" and
     On_Thyroxine_field.get() == "" and
     Query_on_Thyroxine_field.get() == "" and
     On_antithyroid_medication_field.get() == "" and
     Sick_field.get() == "" and
     Pregnant_field.get() == "" and
     Thyroid_Surgery_field.get() == "" and
     T131_treatment_field.get() == "" and
     Query_hypothyroid_field.get() == "" and
     Query_hyperthyroid_field.get() == "" and
     Lithium_field.get() == "" and
     Goitre_field.get() == "" and
     Tumor_field.get() == "" and
     Psych_field.get() == "" and
     TSH_field.get() == "" and
     T3_field.get() == "" and
     TT4_field.get() == "" and
     T4U_field.get() == "" and
     FTI_field.get() == ""  ): 
		  print("empty input") 

   else: 

		# assigning the max row and max column 
		# value upto which data is written 
		# in an excel sheet to the variable 
    current_row = sheet.max_row 
    current_column = sheet.max_column 

		# get method returns current text 
		# as string which we write into 
		# excel spreadsheet at particular location 
    sheet.cell(row=current_row + 1, column=1).value = Age_field.get() 
    sheet.cell(row=current_row + 1, column=2).value = Gender_field.get() 
    sheet.cell(row=current_row + 1, column=3).value = On_Thyroxine_field.get() 
    sheet.cell(row=current_row + 1, column=4).value = Query_on_Thyroxine_field.get() 
    sheet.cell(row=current_row + 1, column=5).value = On_antithyroid_medication_field.get() 
    sheet.cell(row=current_row + 1, column=6).value = Sick_field.get() 
    sheet.cell(row=current_row + 1, column=7).value = Pregnant_field.get()
    sheet.cell(row=current_row + 1, column=8).value = Thyroid_Surgery_field.get()
    sheet.cell(row=current_row + 1, column=9).value = T131_treatment_field.get()
    sheet.cell(row=current_row + 1, column=10).value = Query_hypothyroid_field.get()
    sheet.cell(row=current_row + 1, column=11).value = Query_hyperthyroid_field.get()
    sheet.cell(row=current_row + 1, column=12).value = Lithium_field.get()
    sheet.cell(row=current_row + 1, column=13).value = Goitre_field.get()
    sheet.cell(row=current_row + 1, column=14).value = Tumor_field.get()
    sheet.cell(row=current_row + 1, column=15).value = Psych_field.get()
    sheet.cell(row=current_row + 1, column=16).value = TSH_field.get()
    sheet.cell(row=current_row + 1, column=17).value = T3_field.get()
    sheet.cell(row=current_row + 1, column=18).value = TT4_field.get()
    sheet.cell(row=current_row + 1, column=19).value = T4U_field.get()
    sheet.cell(row=current_row + 1, column=20).value = FTI_field.get()

		# save the file 
    wb.save('D:/Major Project/Thyroid/SEM 8/Python - Sypder/excel.xlsx')
    df = pd.read_excel('D:/Major Project/Thyroid/SEM 8/Python - Sypder/excel.xlsx')  # sheetname is optional
    df.to_csv('D:/Major Project/Thyroid/SEM 8/Python - Sypder/excel.csv', index=False)  # index=False prevents pandas to write row index
    #dataset = pd.read_csv('D:/Major Project/Thyroid/SEM 8/Python - Sypder/excel.csv')

		# set focus on the age_field box 
    Age_field.focus_set() 

		# call the clear() function 
    clear()
    
    final = lda()
    if final==1 :
        #Popup message
        popup("Prediction System", "There are chances of you having thyroid. \n\n For more information refer the link : \n\n https://www.healthline.com/health/hypothyroidism/five-natural-remedies-for-hypothyroidism#1")
    else : 
        #Popup message
        popup("Prediction System", "Congrats! No thyroid detected.\n\n For more information refer the link : \n\n https://www.healthline.com/nutrition/hypothyroidism-diet")

# Driver code 
if __name__ == "__main__": 
	
	# create a GUI window 
   root = tk.Tk() 

	# set the background colour of GUI window 
   root.configure(background='light blue') 

	# set the title of GUI window 
   root.title("Prediction System") 

	# set the configuration of GUI window 
   root.geometry("700x500") 

   excel() 

	# create a Form label 
   heading = tk.Label(root, text="Prediction of Thyroid", bg="light blue") 

	# create labels 
   Age = tk.Label(root, text="Age", bg="light blue") 
   Gender = tk.Label(root, text="Gender", bg="light blue") 
   On_Thyroxine = tk.Label(root, text="On Thyroxine", bg="light blue")  
   Query_on_Thyroxine = tk.Label(root, text="Query on Thyroxine", bg="light blue") 
   On_antithyroid_medication = tk.Label(root, text="On antithyroid medication", bg="light blue")  
   Sick = tk.Label(root, text="Sick", bg="light blue")  
   Pregnant = tk.Label(root, text="Pregnant", bg="light blue") 
   Thyroid_Surgery = tk.Label(root, text="Thyroid Surgery", bg="light blue")
   T131_treatment = tk.Label(root, text="T131 treatment", bg="light blue")
   Query_hypothyroid = tk.Label(root, text="Query hypothyroid", bg="light blue")
   Query_hyperthyroid = tk.Label(root, text="Query hyperthyroid", bg="light blue")
   Lithium = tk.Label(root, text="Lithium", bg="light blue")
   Goitre = tk.Label(root, text="Goitre", bg="light blue")
   Tumor = tk.Label(root, text="Tumor", bg="light blue")
   Psych = tk.Label(root, text="Psych", bg="light blue")
   TSH = tk.Label(root, text="TSH", bg="light blue")
   T3 = tk.Label(root, text="T3", bg="light blue")
   TT4 = tk.Label(root, text="TT4", bg="light blue")
   T4U = tk.Label(root, text="T4U", bg="light blue")
   FTI = tk.Label(root, text="FTI", bg="light blue")

	# grid method is used for placing 
	# the widgets at respective positions 
	# in table like structure . 
   heading.grid(row=0, column=1) 
   Age.grid(row=1, column=0) 
   Gender.grid(row=2, column=0) 
   On_Thyroxine.grid(row=3, column=0) 
   Query_on_Thyroxine.grid(row=4, column=0) 
   On_antithyroid_medication.grid(row=5, column=0) 
   Sick.grid(row=6, column=0) 
   Pregnant.grid(row=7, column=0) 
   Thyroid_Surgery.grid(row=8, column=0)
   T131_treatment.grid(row=9, column=0)
   Query_hypothyroid.grid(row=10, column=0)
   Query_hyperthyroid.grid(row=11, column=0)
   Lithium.grid(row=12, column=0)
   Goitre.grid(row=13, column=0)
   Tumor.grid(row=14, column=0)
   Psych.grid(row=15, column=0)
   TSH.grid(row=16, column=0)
   T3.grid(row=17, column=0)
   TT4.grid(row=18, column=0)
   T4U.grid(row=19, column=0)
   FTI.grid(row=20, column=0)

	# create a text entry box 
	# for typing the information 
   Age_field = tk.Entry(root) 
   Gender_field = tk.Entry(root) 
   On_Thyroxine_field = tk.Entry(root) 
   Query_on_Thyroxine_field = tk.Entry(root) 
   On_antithyroid_medication_field = tk.Entry(root) 
   Sick_field = tk.Entry(root) 
   Pregnant_field = tk.Entry(root)
   Thyroid_Surgery_field = tk.Entry(root)
   T131_treatment_field = tk.Entry(root)
   Query_hypothyroid_field = tk.Entry(root)
   Query_hyperthyroid_field = tk.Entry(root)
   Lithium_field = tk.Entry(root)
   Goitre_field = tk.Entry(root)
   Tumor_field = tk.Entry(root)
   Psych_field = tk.Entry(root)
   TSH_field = tk.Entry(root)
   T3_field = tk.Entry(root)
   TT4_field = tk.Entry(root)
   T4U_field = tk.Entry(root)
   FTI_field = tk.Entry(root)
   
	# bind method of widget is used for 
	# the binding the function with the events 

	# whenever the enter key is pressed 
	# then call the focus2 function 
   Age_field.bind("<Return>", focus1) 
   Gender_field.bind("<Return>", focus2) 

	# whenever the enter key is pressed 
	# then call the focus3 function 
   On_Thyroxine_field.bind("<Return>", focus3) 

	# whenever the enter key is pressed 
	# then call the focus4 function 
   Query_on_Thyroxine_field.bind("<Return>", focus4) 

	# whenever the enter key is pressed 
	# then call the focus5 function 
   On_antithyroid_medication_field.bind("<Return>", focus5) 

	# whenever the enter key is pressed 
	# then call the focus6 function 
   Sick_field.bind("<Return>", focus6) 
   Pregnant_field.bind("<Return>", focus7) 
   Thyroid_Surgery_field.bind("<Return>", focus8) 
   T131_treatment_field.bind("<Return>", focus9) 
   Query_hypothyroid_field.bind("<Return>", focus10) 
   Query_hyperthyroid_field.bind("<Return>", focus11) 
   Lithium_field.bind("<Return>", focus12) 
   Goitre_field.bind("<Return>", focus13) 
   Tumor_field.bind("<Return>", focus14) 
   Psych_field.bind("<Return>", focus15)
   TSH_field.bind("<Return>", focus16) 
   T3_field.bind("<Return>", focus17) 
   TT4_field.bind("<Return>", focus18) 
   T4U_field.bind("<Return>", focus19) 
 

	# grid method is used for placing 
	# the widgets at respective positions 
	# in table like structure . 
   Age_field.grid(row=1,  column=1, ipadx="100") 
   Gender_field.grid(row=2,  column=1, ipadx="100") 
   On_Thyroxine_field.grid(row=3,  column=1, ipadx="100") 
   Query_on_Thyroxine_field.grid(row=4, column=1, ipadx="100") 
   On_antithyroid_medication_field.grid(row=5, column=1, ipadx="100") 
   Sick_field.grid(row=6, column=1, ipadx="100") 
   Pregnant_field.grid(row=7, column=1, ipadx="100") 
   Thyroid_Surgery_field.grid(row=8, column=1, ipadx="100")
   T131_treatment_field.grid(row=9, column=1, ipadx="100")
   Query_hypothyroid_field.grid(row=10, column=1, ipadx="100")
   Query_hyperthyroid_field.grid(row=11, column=1, ipadx="100")
   Lithium_field.grid(row=12, column=1, ipadx="100")
   Goitre_field.grid(row=13, column=1, ipadx="100")
   Tumor_field.grid(row=14, column=1, ipadx="100")
   Psych_field.grid(row=15, column=1, ipadx="100")
   TSH_field.grid(row=16, column=1, ipadx="100")
   T3_field.grid(row=17, column=1, ipadx="100")
   TT4_field.grid(row=18, column=1, ipadx="100")
   T4U_field.grid(row=19, column=1, ipadx="100")
   FTI_field.grid(row=20, column=1, ipadx="100")

	# call excel function 
   excel() 

	# create a Submit Button and place into the root window 
   submit = tk.Button(root, text="Predict", fg="Black", 
							bg="white", command=insert) 

   submit.grid(row=21, column=1) 

	# start the GUI 
   root.mainloop() 
